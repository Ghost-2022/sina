#!/usr/bin/env python
# -*- coding=utf-8 -*-
"""
    date: 2021/10/21 15:43
    author: Ghost
    desc: 
"""
from subprocess import run
import io
import json
import logging
import re
import time
import traceback
import datetime

from openpyxl import Workbook

from common import helper, utils
from libs import router
from config import setting
from modules.sina import SearchHistoryModel, ArticleListModel, CommentListModel,\
    UserModel, LabelRuleModel


@router.Router("/api/v1/qr-cord-url")
class ApiSinaIndexHandler(helper.ApiBaseHandler):
    @utils.login_check
    def get(self, *args, **kwargs):
        session = utils.get_session()
        status = utils.is_login_sina(session)
        if not status:
            qr_id, image = utils.get_qr_code(session)
            logging.info("qr_id: {}, image: {}".format(qr_id, image))
            data = dict(isLogin=status, image=image)
            self.jsonify_finish(is_succ=True, data=data)
            utils.refresh_cookies(session, qr_id)
        else:
            data = dict(isLogin=status)
            self.jsonify_finish(is_succ=True, data=data)


@router.Router("/api/v1/check-login")
class ApiSinaCheckHandler(helper.ApiBaseHandler):
    @utils.login_check
    def get(self, *args, **kwargs):
        session = utils.get_session()
        status = utils.is_login_sina(session)
        data = dict(isLogin=status)
        return self.jsonify_finish(is_succ=True, data=data)


@router.Router("/api/v1/check-spider")
class ApiSinaCheckHandler(helper.ApiBaseHandler):
    @utils.login_check
    def get(self, *args, **kwargs):
        search_id = self.get_argument('searchId', '')
        if not search_id:
            return self.jsonify_finish(error_msg='缺少参数')
        redis_key = setting.SPIDER_STATUS_KEY.format(search_id)
        status = True if self.redis_cache.get(redis_key) else False
        if status:
            self.redis_cache.delete(redis_key)
        data = dict(status=status)
        return self.jsonify_finish(is_succ=True, data=data)


@router.Router("/api/v1/sina-search")
class ApiSinaSearchHandler(helper.ApiBaseHandler):
    @utils.login_check
    def post(self, *args, **kwargs):
        try:
            resp_data = json.loads(self.request.body)
        except Exception:
            logging.error(f'参数异常 {traceback.format_exc()}')
            return self.jsonify_finish(error_msg='系统异常')
        keyword = resp_data.get('keyword')
        start_time = resp_data.get('startTime')
        end_time = resp_data.get('endTime')
        if not all([keyword, start_time, end_time]):
            return self.jsonify_finish(error_msg='缺少参数')
        start_time = datetime.datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
        end_time = datetime.datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
        url_len = self.redis_cache.llen('start_urls')
        if url_len > 0:
            return self.jsonify_finish(error_msg=u'已有程序正在执行，请稍后再试')
        cursor, conn = self.application.db_pool.get_conn()
        data = {'isDownloading': False}
        try:
            row_id = SearchHistoryModel.insert_record(
                keyword, start_time, end_time, cursor)
        except Exception:
            logging.error(f'数据插入失败{traceback.format_exc()}')
            return self.jsonify_finish(error_msg=u'系统繁忙')
        else:
            json_data = {
                'start_time': int(start_time.timestamp()),
                'end_time': int(end_time.timestamp()),
                'keyword': keyword,
                'search_id': row_id
            }
            self.redis_cache.rpush('start_urls', json.dumps(json_data))
            data['isDownloading'] = True
            data['searchId'] = row_id
            return self.jsonify_finish(is_succ=True, data=data)

    @utils.login_check
    def get(self):
        search_id = self.get_argument('searchId', '')
        if not search_id:
            return self.jsonify_finish(error_msg='缺少参数：searchId')
        cursor, conn = self.application.db_pool.get_conn()
        condition = {
            'id': search_id,
        }
        record = SearchHistoryModel.get_record(condition, cursor)
        if not record:
            return self.jsonify_finish(error_msg=u'数据不存在')
        search_id, info = record['id'], record['info']
        try:
            info = json.loads(info)
        except Exception:
            logging.error(f'info 解析失败：\n{traceback.format_exc()}')
            info = {}
        article_data = ArticleListModel.get_data_group_by_date(
            search_id, cursor)
        comment_data = CommentListModel.get_data_group_by_date(
            search_id, cursor)
        data = {
            'commentCounts': info.get('comment_counts', []),
            'articleCounts': info.get('article_counts', []),
            'articleEmotion': info.get('article_emotion', []),
            'commentEmotion': info.get('comment_emotion', []),
            'articleData': article_data,
            'commentData': comment_data,
            'articleGroup': info.get('a_group_count', []),
            'commentGroup': info.get('c_group_count', []),
            'articleCloud': f'static/search_{search_id}/article.jpg',
            'commentCloud': f'static/search_{search_id}/comment.jpg'}
        self.jsonify_finish(is_succ=True, data=data)

    @utils.login_check
    def delete(self):
        search_id = self.get_argument('searchId', '')
        if not search_id:
            return self.jsonify_finish(error_msg='缺少参数：searchId')
        cursor, conn = self.application.db_pool.get_conn()
        result = SearchHistoryModel.drop_record(search_id, cursor)
        if result:
            self.jsonify_finish(is_succ=True, error_msg='删除成功')
        else:
            self.jsonify_finish(is_succ=True, error_msg='数据不存在')


@router.Router('/api/v1/search-list')
class SearchListHandler(helper.ApiBaseHandler):
    @utils.login_check
    def get(self):
        page = self.get_argument('page', '')
        size = self.get_argument('size', '10')
        status = self.get_argument('status', '1')
        page = int(page) if page.isdigit() else 1
        cursor, conn = self.application.db_pool.get_conn()
        condition = {
            'status': status,
        }
        records = SearchHistoryModel.get_records(
            condition, cursor, offset=page - 1, limit=size)
        count = SearchHistoryModel.count_records(condition, cursor)
        records_data = [{
            'id': item['id'],
            'keyword': item['keyword'],
            'startTime': item['start_time'].strftime('%Y-%m-%d %H:%M:%S'),
            'endTime': item['end_time'].strftime('%Y-%m-%d %H:%M:%S')}
            for item in records]
        data = {
            'list': records_data,
            'total': count
        }
        return self.jsonify_finish(is_succ=True, data=data)


@router.Router('/api/v1/get-token')
class TokenHandler(helper.ApiBaseHandler):
    def post(self):
        try:
            data = json.loads(self.request.body)
        except Exception :
            logging.error(f'参数解析失败：{traceback.format_exc()}')
            return self.jsonify_finish(error_msg=u'参数异常')
        username = data.get('username', '')
        password = data.get('password', '')
        secret = setting.SECRET_KEY
        encry_pwd = utils.encrypt_hamc_sha256(secret, password)
        cursor, conn = self.application.db_pool.get_conn()
        user = UserModel.get_user(username, encry_pwd, cursor)
        data = {
            'token': '',
            'refreshToken': '',
            'expiration': 0,
            'nickName': ''
        }
        # logging.info(f'username:{username}, pwd:{password}, user:{user}')
        if user:
            exp = int(time.time()) + 3600 * 24
            token, refresh_token = utils.create_token(
                user['id'], user['username'], exp)
            data['expiration'] = exp
            data['token'] = token
            data['refreshToken'] = refresh_token
            data['nickName'] = username
            return self.jsonify_finish(is_succ=True, data=data)
        return self.jsonify_finish(error_msg='验证失败')


@router.Router('/api/v1/refresh-token')
class RefreshTokenHandler(helper.ApiBaseHandler):

    @utils.refresh_token
    def post(self):
        if 'Authorization' in self.request.headers:
            token = self.request.headers.get('Authorization').split()[-1]
        else:
            token = self.get_argument('token', '')

        status, data = utils.verify_refresh_token(token)
        user_id = data.get('user_id')
        username = data.get('username')
        exp = int(time.time()) + 3600 * 24
        token, refresh_token = utils.create_token(user_id, username, exp)
        data = {
            'token': token,
            'refreshToken': refresh_token,
            'expiration': exp
        }
        return self.jsonify_finish(is_succ=True, data=data)


@router.Router('/api/v1/user')
class UserHandler(helper.ApiBaseHandler):
    def post(self):
        username = self.get_argument('username', '')
        password = self.get_argument('password', '')
        if not all([username, password]):
            return self.jsonify_finish(error_msg=u'参数错误')
        secret = setting.SECRET_KEY
        encry_pwd = utils.encrypt_hamc_sha256(secret, password)
        cursor, conn = self.application.db_pool.get_conn()
        try:
            UserModel.create_user(username, encry_pwd, cursor)
            conn.commit()
        except Exception:
            logging.error(f'创建用户失败 \n {traceback.format_exc()}')
            return self.jsonify_finish(error_msg='系统繁忙')
        else:
            return self.jsonify_finish(is_succ=True, error_msg='添加成功')


@router.Router('/api/v1/export-article')
class ExportArticleHandler(helper.ApiBaseHandler):
    def get(self):
        search_id = self.get_argument('searchId')
        cursor, conn = self.application.db_pool.get_conn()
        article_list = ArticleListModel.query_records_by_search_id(search_id, cursor)

        article_data = [[
            '用户名', '用户主页（网址）', '微博内容', '转发', '评论', '点赞', '时间',
            '微博链接', '经度', '维度'
        ]]
        for item in article_list:
            article_data.append([
                item['author'],
                item['author_url'],
                re.sub('[\n]*?', '', item['content']),
                item['reposts_count'],
                item['comments_count'],
                item['attitudes_count'],
                item['publish_time'].strftime('%Y-%m-%d %H:%M:%S'),
                item['article_url'],
                item['lng'],
                item['lat']
            ])
        utils.export_to_csv(self, '{}-article.csv'.format(search_id), article_data)


@router.Router('/api/v1/export-comment')
class ExportCommentHandler(helper.ApiBaseHandler):

    def get(self):
        search_id = self.get_argument('searchId')
        cursor, conn = self.application.db_pool.get_conn()
        comment_list = CommentListModel.query_records_by_search_id(search_id, cursor)

        comment_data = [[
            '用户名',  '用户链接', '评论内容', '时间', '点赞', '微博链接'
        ]]
        for item in comment_list:
            comment_data.append([
                item['author'],
                item['author_url'],
                re.sub('[\n]*?', '', item['content']),
                item['publish_time'].strftime('%Y-%m-%d %H:%M:%S'),
                item['like_counts'],
                item['article_url'],
            ])
        utils.export_to_csv(self, '{}-comment.csv'.format(search_id), comment_data)


@router.Router('/api/v1/label-rule')
class LabelRuleHandler(helper.ApiBaseHandler):
    @utils.login_check
    def get(self):
        page = int(self.get_argument('page', '1'))
        page_size = int(self.get_argument('pageSize', '10'))
        label = self.get_argument('label', '')
        cursor, conn = self.application.db_pool.get_conn()
        condition = {}
        if label:
            condition['label'] = label
        base_data = LabelRuleModel.get_labels(condition, cursor, offset=(page-1)*page_size, limit=page_size)
        count = LabelRuleModel.count_total_label(condition, cursor)
        data = {
            'list': base_data,
            'total': count
        }
        return self.jsonify_finish(is_succ=True, error_msg=u'', data=data)

    @utils.login_check
    def post(self):
        try:
            data = json.loads(self.request.body)
        except Exception:
            logging.error(f'参数解析失败：{traceback.format_exc()}')
            return self.jsonify_finish(error_msg=u'参数异常')
        label = data.get('label', '')
        rule = data.get('rule', '')
        label_id = data.get('labelId', '')
        if not all([label, rule]):
            return self.jsonify_finish(error_msg=u'参数错误')
        cursor, conn = self.application.db_pool.get_conn()
        try:
            if not label_id:
                LabelRuleModel.insert_label(label, rule, cursor)
            else:
                LabelRuleModel.update_label(label_id, label, rule, cursor)
        except Exception:
            logging.error(f'规则添加失败 {traceback.format_exc()}')
            return self.jsonify_finish(error_msg=u'系统繁忙')
        else:
            return self.jsonify_finish(is_succ=True, error_msg=u'添加成功')

    @utils.login_check
    def delete(self):
        try:
            data = json.loads(self.request.body)
        except Exception:
            logging.error(f'参数解析失败：{traceback.format_exc()}')
            return self.jsonify_finish(error_msg=u'参数异常')
        label_id = data.get('labelIds')
        if not label_id:
            return self.jsonify_finish(error_msg=u'缺少参数')
        label_ids = label_id.split(',')
        cursor, conn = self.application.db_pool.get_conn()
        try:
            LabelRuleModel.del_label(label_ids, cursor)
        except Exception:
            logging.error(f'label {label_id} 删除失败， {traceback.format_exc()}')
            return self.jsonify_finish(error_msg=u'系统繁忙')
        else:
            return self.jsonify_finish(is_succ=True)


@router.Router('/api/v1/get-points')
class MapPointsHandler(helper.ApiBaseHandler):
    @utils.login_check
    def get(self):
        search_id = self.get_argument('searchId')
        query_type = self.get_argument('type', 'article')
        if not search_id:
            return self.jsonify_finish(error_msg=u'缺少参数')
        cursor, conn = self.application.db_pool.get_conn()
        if query_type == 'article':
            base_data = ArticleListModel.query_points_by_search_id(search_id, cursor)
        else:
            base_data = CommentListModel.query_points_by_search_id(search_id, cursor)
        data = [[item['lng'], item['lat']]for item in base_data]
        return self.jsonify_finish(is_succ=True, data=data)


@router.Router('/api/v1/export-article-cate')
class ExportArticleHandler(helper.ApiBaseHandler):

    def get(self):
        search_id = self.get_argument('searchId')
        cursor, conn = self.application.db_pool.get_conn()
        article_list = ArticleListModel.query_records_by_search_id(search_id, cursor)
        titles = ['用户名', '用户主页（网址）', '微博内容', '转发', '评论', '点赞', '时间',
                  '微博链接', '经度', '维度']
        data = {}
        for item in article_list:
            for cate in item['cate_list'].split(','):
                data.setdefault(cate, [titles]).append([
                    item['author'],
                    item['author_url'],
                    re.sub('[\n]*?', '', item['content']),
                    item['reposts_count'],
                    item['comments_count'],
                    item['attitudes_count'],
                    item['publish_time'].strftime('%Y-%m-%d %H:%M:%S'),
                    item['article_url'],
                    item['lng'],
                    item['lat']
                ])
        std = io.BytesIO()
        wb = Workbook()
        for title, content in data.items():
            w_sheet = wb.create_sheet(title)
            for text in content:
                w_sheet.append(text)
        wb.remove(wb[wb.sheetnames[0]])
        wb.save(std)
        # http头 浏览器自动识别为文件下载
        self.set_header('Content-Type', 'application/octet-stream')
        # 下载时显示的文件名称
        self.set_header('Content-Disposition',
                        'attachment; filename={0}-article-cate.xlsx'.format(search_id))
        self.write(std.getvalue())
        return self.finish()


@router.Router('/api/v1/export-comment-cate')
class ExportArticleHandler(helper.ApiBaseHandler):

    def get(self):
        search_id = self.get_argument('searchId')
        cursor, conn = self.application.db_pool.get_conn()
        article_list = ArticleListModel.query_records_by_search_id(search_id, cursor)
        titles = ['用户名',  '用户链接', '评论内容', '时间', '点赞', '微博链接']
        data = {}
        for item in article_list:
            for cate in item['cate_list'].split(','):
                data.setdefault(cate, [titles]).append([
                item['author'],
                item['author_url'],
                re.sub('[\n]*?', '', item['content']),
                item['publish_time'].strftime('%Y-%m-%d %H:%M:%S'),
                item['like_counts'],
                item['article_url'],
            ])
        std = io.BytesIO()
        wb = Workbook()
        for title, content in data.items():
            w_sheet = wb.create_sheet(title)
            for text in content:
                w_sheet.append(text)
        wb.remove(wb[wb.sheetnames[0]])
        wb.save(std)
        # http头 浏览器自动识别为文件下载
        self.set_header('Content-Type', 'application/octet-stream')
        # 下载时显示的文件名称
        self.set_header('Content-Disposition',
                        'attachment; filename={0}-comment-cate.xlsx'.format(search_id))
        self.write(std.getvalue())
        return self.finish()